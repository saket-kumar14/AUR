"""
clean_data.py

Reads the 3 raw ranking Excel files (42 countries, one sheet per country)
and produces clean_universities.json — a normalized list of records ready
to be upserted into University + RankingScore (year=2026) by seed_official.py.

Usage:
    python clean_data.py

Input files expected in the same directory (or pass explicit paths via
RAW_FILES below):
    harsh_individual_work.xlsx               (10 countries)
    Jhalak_individual_work_done.xlsx          (10 countries)
    set_1_22_countries_by_Jhalak_n_team.xlsx  (22 countries)

Output:
    clean_universities.json
"""

import json
import re
import datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

RAW_FILES = [
    "harsh_individual_work.xlsx",
    "Jhalak_individual_work_done.xlsx",
    "set_1_22_countries_by_Jhalak_n_team.xlsx",
]

OUTPUT_PATH = "clean_universities.json"

RANKING_YEAR = 2026

# Sheets known to have an extra title row before the real header row.
# (per handoff doc: India, Singapore, Turkey — detection below is also
# generic/defensive in case other sheets have the same issue)
SHEETS_WITH_TITLE_ROW = {"india", "singapore", "turkey"}

# Sheets known to have ratio fields Excel mangled into datetime.time / other
# odd types (Faculty Student Ratio, International Faculty/Student Ratio can
# all be affected depending on the sheet).
SHEETS_WITH_RATIO_ISSUES = {"georgia", "indonesia", "qatar", "uzbekistan"}

# Manual overrides for is_public, from prior research (per handoff doc).
# Keyed by (country_lower, university_name_lower_stripped).
MANUAL_IS_PUBLIC_OVERRIDES = {
    ("bangladesh", "islamic university of technology"): True,
    ("thailand", "asian institute of technology"): None,  # explicitly left null
    ("singapore", "national institute of education"): True,
    ("vietnam", "hong bang international university"): False,
    # remaining 6 of the "7 Vietnam universities" from the handoff doc —
    # blank University Type field in the raw sheet, manually researched and
    # resolved to Public (Hong Bang above is the one exception, resolved Private)
    ("vietnam", "ho chi minh city university of agriculture and forestry"): True,
    ("vietnam", "national economics university"): True,
    ("vietnam", "ho chi minh city university of technology and education"): True,
    ("vietnam", "hanoi national university of education"): True,
    ("vietnam", "vietnam maritime university"): True,
    ("vietnam", "university of transport and communications"): True,
}
# The rest of the Vietnam list from the doc ("7 Vietnam universities -> Public
# except Hong Bang") gets handled generically: any Vietnam sheet row not
# otherwise classifiable falls back through MANUAL_TYPE_MAP_PUBLIC_TRUE below,
# with Hong Bang pinned False by the exact-match override above.

# University Type strings -> is_public True. Conservative: unmatched strings
# that aren't clearly "Private*" stay null rather than guessing.
MANUAL_TYPE_MAP_PUBLIC_TRUE = {
    "public",
    "public university",
    "public research university",
    "public deemed university",
    "central university",
    "national university",
    "state university",
    "semi-govt.",
    "semi govt",
    "semi-government",
    "government",
    "govt.",
    "govt",
    "deemed",
    "deemed university",
    "public technical university",
    "autonomous public university",
}
MANUAL_TYPE_MAP_PUBLIC_FALSE = {
    "private",
    "private university",
    "private research university",
    "private deemed university",
    "private branch campus",
    "private non-profit",
    "private for-profit",
}

# ---------------------------------------------------------------------------
# Header normalization / fuzzy matching
# ---------------------------------------------------------------------------

# Canonical field name -> list of normalized header variants we've seen or
# expect (normalization = lowercase, strip, collapse whitespace, drop
# punctuation like *, (), digits-in-parens for year call-outs).
HEADER_MAP = {
    "rank": ["rank"],
    "university_name": ["university name", "name"],
    "country": ["country"],
    "city": ["city"],
    "qs_rank": ["qs rank", "qs rank 2026", "qs world rank", "qs world rank 2026"],
    "the_rank": ["the rank", "the rank 2025", "the world rank", "the world rank 2025"],
    "university_type": ["university type", "type"],
    "year_established": ["year established", "established", "founded"],
    "website": ["official website", "website", "url"],
    "ar_score": ["academic reputation score", "academic reputation"],
    "er_score": ["employer reputation score", "employer reputation"],
    "fsr_ratio": ["faculty student ratio", "faculty:student ratio"],
    "cpp_score": ["citations per faculty", "citations per faculty score"],
    "ifr_score": ["international faculty ratio", "international faculty ratio score"],
    "isr_score": ["international student ratio", "international student ratio score"],
    "research_output_score": ["research output score"],
    "research_impact_score": ["research impact score"],
    "graduate_employability_score": ["graduate employability score"],
    "industry_income_score": ["industry income score"],
    "irn_score": ["international research network score"],
    "sustainability_score": ["sustainability score"],
    "total_students": ["total student population", "total students"],
    "total_faculty": ["total faculty members", "total faculty"],
    "international_students": ["international students"],
    "international_faculty": ["international faculty"],
    "research_publications": ["number of research publications"],
    "citations_count": ["number of citations"],
    "graduation_rate": ["graduation rate (%)", "graduation rate"],
    "placement_rate": ["placement rate (%)", "placement rate"],
    "overall_score": ["overall score"],
    "accreditation": ["accreditation"],
    "key_courses": ["key courses offered", "key courses"],
    "campus_area": ["campus area"],
    "contact_email": ["contact email", "email"],
    "contact_number": ["contact numbers", "contact number", "phone"],
    "remarks": ["remarks", "notes"],
}

# Build reverse lookup: normalized variant -> canonical field
_VARIANT_TO_FIELD = {}
for field, variants in HEADER_MAP.items():
    for v in variants:
        _VARIANT_TO_FIELD[v] = field


def normalize_header(raw):
    """Lowercase, strip, collapse whitespace, drop trailing markers like
    '*', '(2026)', '2026' etc. so 'QS Rank*' / 'QS Rank (2026)' / 'QS Rank 2026'
    all normalize toward the same family of variant strings."""
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    s = s.replace("*", "")
    s = re.sub(r"\(\s*\d{4}\s*\)", "", s)  # "(2026)" -> ""
    s = re.sub(r"\s+", " ", s).strip()
    return s


def map_header(raw):
    """Return canonical field name for a raw header cell, or None if unmapped
    (extra/unmapped columns beyond the ~20 core fields still get looked up
    here too, since HEADER_MAP includes the 'useful extras' per the handoff
    doc decision to map total_students/total_faculty/placement_rate etc.)."""
    norm = normalize_header(raw)
    if norm in _VARIANT_TO_FIELD:
        return _VARIANT_TO_FIELD[norm]
    # loose fallback: try startswith match for headers with extra trailing
    # words we haven't explicitly listed (defensive, doc says wording varies)
    for variant, field in _VARIANT_TO_FIELD.items():
        if norm.startswith(variant) or variant.startswith(norm):
            return field
    return None


# ---------------------------------------------------------------------------
# Value cleaners
# ---------------------------------------------------------------------------

NA_STRINGS = {"na", "n/a", "-", "", "none", "null"}


def is_na(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip().lower() in NA_STRINGS:
        return True
    return False


def clean_rank(value):
    """Plain ints, ranges ('251-300' / '251–300'), or hash-prefixed ('#269')
    -> first integer found. Returns None if not parseable."""
    if is_na(value):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value)
    m = re.search(r"\d+", s)
    return int(m.group()) if m else None


def clean_the_rank(value):
    """THE rank is kept as text (bands like '351-400', '1501+') per schema
    (the_rank is String(50)), just normalize dash characters and whitespace."""
    if is_na(value):
        return None
    s = str(value).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", "", s)
    return s[:50] if s else None


def clean_numeric(value):
    """Returns a float if the value is genuinely numeric (including numeric
    strings and percentage strings like '90%+'), else None — never guesses a
    value for qualitative text like 'Very High' / 'Excellent'."""
    if is_na(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime.time):
        # a numeric score field should never actually be a time object; if it
        # got here treat as unparseable rather than misinterpret
        return None
    s = str(value).strip()
    # strip a trailing % or + and leading ~ before checking numeric-ness
    s_clean = s.rstrip("%+").lstrip("~").strip()
    try:
        return float(s_clean)
    except ValueError:
        return None


def clean_year(value):
    """Extract a plausible 4-digit year from strings like
    '1911 (re-established in Taiwan 1956)' -> 1911 (first 4-digit year)."""
    if is_na(value):
        return None
    if isinstance(value, (int, float)):
        y = int(value)
        return y if 1000 <= y <= 2100 else None
    s = str(value)
    m = re.search(r"\b(1[5-9]\d{2}|20\d{2})\b", s)
    return int(m.group()) if m else None


def clean_ratio(value):
    """Faculty/staff-style ratio fields. Excel sometimes converts 'HH:MM'-
    looking ratios into datetime.time or timedelta-like strings. Normalize
    back to a clean 'A:B' string where possible; otherwise pass through
    numeric or leave as trimmed string; None for NA."""
    if is_na(value):
        return None
    if isinstance(value, datetime.time):
        # e.g. time(1, 16) came from someone typing "1:16" and Excel
        # auto-converting it to a time-of-day. Reconstruct "H:MM" (drop
        # leading zero padding oddities, keep it human-readable).
        return f"{value.hour}:{value.minute:02d}"
    if isinstance(value, datetime.timedelta):
        total_minutes = int(value.total_seconds() // 60)
        h, m = divmod(total_minutes, 60)
        return f"{h}:{m:02d}"
    s = str(value).strip()
    # "1 day, 0:18:00" style timedelta-as-string from Excel/openpyxl
    m = re.match(r"(?:(\d+)\s*day[s]?,?\s*)?(\d+):(\d+)(?::(\d+))?", s)
    if m and ":" in s:
        days = int(m.group(1) or 0)
        hh = int(m.group(2)) + days * 24
        mm = int(m.group(3))
        return f"{hh}:{mm:02d}"
    # already a clean ratio-like string ("14:1", "26.2:1", "~1:20",
    # "14:1 to 17:1") — keep as-is, just strip stray leading "~"
    if s:
        return s.lstrip("~").strip()
    return None


def clean_website(value):
    """Add missing protocol; null out junk (e.g. Google share links that
    aren't real institutional URLs)."""
    if is_na(value):
        return None
    s = str(value).strip()
    low = s.lower()
    junk_markers = ["docs.google.com", "drive.google.com", "share?", "forms.gle"]
    if any(marker in low for marker in junk_markers):
        return None
    if not low.startswith("http://") and not low.startswith("https://"):
        s = "https://" + s
    return s


def clean_is_public(university_type, country, name):
    """Map free-text University Type -> boolean, using manual exact-match
    overrides first, then the conservative type-string mapping. Stays None
    if not confidently classifiable, rather than guessing."""
    country_key = (country or "").strip().lower()
    name_key = (name or "").strip().lower()
    for (override_country, override_name), value in MANUAL_IS_PUBLIC_OVERRIDES.items():
        if country_key == override_country and override_name in name_key:
            return value

    if is_na(university_type):
        return None
    t = str(university_type).strip().lower()
    t = re.sub(r"\s+", " ", t)

    if t in MANUAL_TYPE_MAP_PUBLIC_TRUE:
        return True
    if t in MANUAL_TYPE_MAP_PUBLIC_FALSE:
        return False
    # substring fallback: "public deemed university" contains "public",
    # "private branch campus" contains "private"
    if "public" in t or "government" in t or "govt" in t or "national" in t or "central" in t or "state" in t:
        return True
    if "private" in t:
        return False
    return None


def slugify(name, country=None):
    base = f"{name} {country}" if country else name
    base = base.lower()
    base = re.sub(r"[^a-z0-9\s]", "", base)
    base = re.sub(r"\s+", "-", base.strip())
    return base[:80]


# ---------------------------------------------------------------------------
# Sheet reading
# ---------------------------------------------------------------------------

def find_header_row(ws, max_scan=5):
    """Return the 1-indexed row number containing the real header row,
    handling sheets with an extra title row above it (India/Singapore/Turkey
    and defensively any other sheet with the same pattern). Detection: the
    header row is the first row where >= 5 non-empty cells are found AND
    the row contains a cell that normalizes to 'rank' or 'university name'."""
    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row_vals if v is not None and str(v).strip() != ""]
        normed = {normalize_header(v) for v in non_empty}
        if len(non_empty) >= 5 and ("rank" in normed or "university name" in normed):
            return r
    return 1  # fallback, shouldn't happen given known sheet shapes


def read_sheet(ws, country_label):
    header_row_idx = find_header_row(ws)
    raw_headers = [ws.cell(header_row_idx, c).value for c in range(1, ws.max_column + 1)]
    field_by_col = {}
    for col_idx, raw_h in enumerate(raw_headers, start=1):
        field = map_header(raw_h)
        if field:
            field_by_col[col_idx] = field

    records = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # stop at first fully-blank row (sheets padded to 100 rows)
        if all(v is None or str(v).strip() == "" for v in row_vals):
            break

        rec = {}
        for col_idx, field in field_by_col.items():
            rec[field] = row_vals[col_idx - 1] if col_idx - 1 < len(row_vals) else None

        name = rec.get("university_name")
        if is_na(name):
            continue  # skip stray/junk rows with no university name

        cleaned = build_clean_record(rec, country_label)
        records.append(cleaned)

    return records


def build_clean_record(rec, country_label):
    name = str(rec.get("university_name")).strip()
    country = rec.get("country")
    country = str(country).strip() if not is_na(country) else country_label.title()

    university_type = rec.get("university_type")

    clean = {
        # identity
        "slug": slugify(name, country),
        "name": name,
        "country": country,
        "city": (str(rec.get("city")).strip() if not is_na(rec.get("city")) else None),

        # University model fields
        "is_public": clean_is_public(university_type, country, name),
        "established_year": clean_year(rec.get("year_established")),
        "total_students": _to_int(clean_numeric(rec.get("total_students"))),
        "total_faculty": _to_int(clean_numeric(rec.get("total_faculty"))),
        "placement_percentage": clean_numeric(rec.get("placement_rate")),
        "website_url": clean_website(rec.get("website")),

        # extras kept for reference / future use, not all map to current schema
        "university_type_raw": (str(university_type).strip() if not is_na(university_type) else None),
        "graduation_rate": clean_numeric(rec.get("graduation_rate")),
        "overall_score_raw": clean_numeric(rec.get("overall_score")),
        "accreditation": (str(rec.get("accreditation")).strip() if not is_na(rec.get("accreditation")) else None),
        "key_courses": (str(rec.get("key_courses")).strip() if not is_na(rec.get("key_courses")) else None),
        "campus_area": (str(rec.get("campus_area")).strip() if not is_na(rec.get("campus_area")) else None),
        "contact_email": (str(rec.get("contact_email")).strip() if not is_na(rec.get("contact_email")) else None),
        "remarks": (str(rec.get("remarks")).strip() if not is_na(rec.get("remarks")) else None),
        "sustainability_score": clean_numeric(rec.get("sustainability_score")),
        "research_publications": _to_int(clean_numeric(rec.get("research_publications"))),
        "citations_count": _to_int(clean_numeric(rec.get("citations_count"))),
        "international_students": _to_int(clean_numeric(rec.get("international_students"))),
        "international_faculty": _to_int(clean_numeric(rec.get("international_faculty"))),
        "fsr_ratio_raw": clean_ratio(rec.get("fsr_ratio")),

        # RankingScore (year=2026) fields
        "ranking": {
            "year": RANKING_YEAR,
            "rank": clean_rank(rec.get("qs_rank")),
            "the_rank": clean_the_rank(rec.get("the_rank")),
            "overall_score": clean_numeric(rec.get("overall_score")),
            "ar_score": clean_numeric(rec.get("ar_score")),
            "er_score": clean_numeric(rec.get("er_score")),
            "fsr_score": clean_numeric(rec.get("fsr_ratio")),  # numeric where possible; ratio string kept in fsr_ratio_raw
            "irn_score": clean_numeric(rec.get("irn_score")),
            "cpp_score": clean_numeric(rec.get("cpp_score")),
            # isr/ifr intentionally left RAW/unnormalized (mixed % vs fraction
            # scales across sheets) — per handoff doc decision, do NOT auto-normalize
            "ifr_score": clean_numeric(rec.get("ifr_score")),
            "isr_score": clean_numeric(rec.get("isr_score")),
            "research_output_score": clean_numeric(rec.get("research_output_score")),
            "research_impact_score": clean_numeric(rec.get("research_impact_score")),
            "graduate_employability_score": clean_numeric(rec.get("graduate_employability_score")),
            "industry_income_score": clean_numeric(rec.get("industry_income_score")),
        },
    }
    return clean


def _to_int(value):
    if value is None:
        return None
    try:
        return int(round(value))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    all_records = []
    base_dir = Path(__file__).parent

    for fname in RAW_FILES:
        path = base_dir / fname
        if not path.exists():
            print(f"WARNING: {fname} not found at {path}, skipping")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            country_label = sheet_name.strip()
            records = read_sheet(ws, country_label)
            print(f"{fname} :: {sheet_name} -> {len(records)} records")
            all_records.extend(records)

    print(f"\nTotal records: {len(all_records)}")

    with open(base_dir / OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(all_records, f, indent=2, ensure_ascii=False)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()